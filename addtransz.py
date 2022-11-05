# cython: language_level=3
# !/usr/bin/python
# -*- coding: UTF-8 -*-
# @Date    : 2022-10-25 00:23:45
# @Function: Auto-drawing single line diagram
# @Author  : Luo Xiaochun
# @Email   : luoxiaochun@proton.me
# @version : 1.1.0

import sqlite3
import settings
import win32api
import win32con
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import numpy as np
import math
from single_line import polemodify


def addtransz(transzname):
    def intput_polepath(inpolepath):
        polepath_raw = inpolepath
        # polepath_raw = 'X12.F7.8.H6.Z1'
        polesplit = polepath_raw.split('.')
        for j in range(len(polesplit)):
            if 'F' in polesplit[j]:
                polesplit[j] = (polesplit[j].replace('F', 'F&'))
            if 'H' in polesplit[j]:
                polesplit[j] = (polesplit[j].replace('H', 'H&'))
            if 'X' in polesplit[j]:
                polesplit[j] = (polesplit[j].replace('X', 'X&'))
            if 'Z' in polesplit[j]:
                polesplit[j] = (polesplit[j].replace('Z', 'Z&'))
            if 'Y' in polesplit[j]:
                polesplit[j] = (polesplit[j].replace('Y', 'Y&'))
            if 'L' in polesplit[j]:
                polesplit[j] = (polesplit[j].replace('L', 'L&'))

            if 'f' in polesplit[j]:
                polesplit[j] = (polesplit[j].replace('f', 'F&'))
            if 'h' in polesplit[j]:
                polesplit[j] = (polesplit[j].replace('h', 'H&'))
            if 'x' in polesplit[j]:
                polesplit[j] = (polesplit[j].replace('x', 'X&'))
            if 'z' in polesplit[j]:
                polesplit[j] = (polesplit[j].replace('z', 'Z&'))
            if 'y' in polesplit[j]:
                polesplit[j] = (polesplit[j].replace('y', 'Y&'))
            if 'l' in polesplit[j]:
                polesplit[j] = (polesplit[j].replace('l', 'L&'))

        inpolepath = '.'.join(polesplit)
        return inpolepath

    # inpoleact = intput_polepath('12.F7.8.H6.1')

    # print('inpoleact =', inpoleact )

    def output_polepath(inpolepath):
        polepath_raw = inpolepath
        # polepath_raw = 'X&12.F&7.8.H&6.Z&1'
        if polepath_raw:
            polesplit = polepath_raw.split('.')
        else:
            polepath_raw = '0.0'
            polesplit = polepath_raw.split('.')
        for j in range(len(polesplit)):
            if 'F' in polesplit[j] or 'H' in polesplit[j] or 'X' in polesplit[j] or 'Z' in polesplit[j] or 'Y' in \
                    polesplit[
                        j] or 'L' in polesplit[j]:
                polesplit[j] = polesplit[j].replace('&', '')
            if 'f' in polesplit[j] or 'h' in polesplit[j] or 'x' in polesplit[j] or 'z' in polesplit[j] or 'y' in \
                    polesplit[
                        j] or 'l' in polesplit[j]:
                polesplit[j] = polesplit[j].replace('&', '')
        outpolepath = '.'.join(polesplit)
        return outpolepath

    transz_name1 = transzname

    # transz_name = 'inputdata' + '\\' + '\\' + str(transz_name1)

    transz_name = str(transz_name1)

    l_name = settings.line_names
    para_name = settings.lpara_names


    line_tpath = 'LGJ-120.LGJ-95.LGJ-70'
    line_lpath = '1250.750.100'

    print('transz_name=', transz_name)
    # --------------------导入新数据Start----------------------------

    # -------------------配变坐标参数插入数据库rawdata----------------
    conn = sqlite3.connect("rawdata.db3")
    cut = conn.cursor()
    try:
        transz_para = []
        wb = load_workbook(transz_name)
        ws = wb['transz']

        for row_A in range(2, ws.max_row + 1, 1):
            a1 = ws.cell(row=row_A, column=1).value
            a2 = ws.cell(row=row_A, column=2).value
            a3 = ws.cell(row=row_A, column=3).value
            a4 = ws.cell(row=row_A, column=4).value
            a5 = ws.cell(row=row_A, column=5).value
            a6 = ws.cell(row=row_A, column=6).value
            a7 = ws.cell(row=row_A, column=7).value
            a8 = ws.cell(row=row_A, column=8).value
            a9 = ws.cell(row=row_A, column=9).value
            a10 = ws.cell(row=row_A, column=10).value
            a11 = ws.cell(row=row_A, column=11).value
            a12 = ws.cell(row=row_A, column=12).value
            a13 = ws.cell(row=row_A, column=13).value
            a14 = ws.cell(row=row_A, column=14).value
            a15 = ws.cell(row=row_A, column=15).value
            a16 = ws.cell(row=row_A, column=16).value
            a17 = ws.cell(row=row_A, column=17).value
            a18 = ws.cell(row=row_A, column=18).value
            a19 = ws.cell(row=row_A, column=19).value
            a20 = ws.cell(row=row_A, column=20).value
            a21 = ws.cell(row=row_A, column=21).value
            a22 = ws.cell(row=row_A, column=22).value
            a23 = ws.cell(row=row_A, column=23).value
            a24 = ws.cell(row=row_A, column=24).value
            a25 = ws.cell(row=row_A, column=25).value
            a26 = ws.cell(row=row_A, column=26).value

            a27 = ws.cell(row=row_A, column=27).value
            a28 = ws.cell(row=row_A, column=28).value
            a29 = ws.cell(row=row_A, column=29).value
            a30 = ws.cell(row=row_A, column=30).value
            a31 = ws.cell(row=row_A, column=31).value
            a32 = ws.cell(row=row_A, column=32).value
            a33 = ws.cell(row=row_A, column=33).value

            a34 = ws.cell(row=row_A, column=34).value

            # -------------
            if a25:
                a25 = a25
            else:
                a25 = 0.0
            if a26:
                a26 = a26
            else:
                a26 = 0.0
            if a27:
                a27 = a27
            else:
                a27 = 0.0
            if a28:
                a28 = a28
            else:
                a28 = 0.0
            if a32:
                a32 = a32
            else:
                a32 = a10

            # -------------
            if a4 != '' and a7 != '' and a10 != '':
                transz_para.append(
                    {'id': a1, 'tnode': a2, 'company': a3, 'tname': a4, 'tnumber': a5, 'tcapacity': a6, 'lname': a7,
                     'tpath': a8,
                     'lpath': a9, 'polepath': a10, 'rline': a11, 'xline': a12, 'zline': a13, 'deltaline': a14,
                     'allline_length': a15,
                     'rt': a16, 'xt': a17, 'zt': a18, 'deltazt': a19, 'ztsecond': a20, 'ldeltaall': a21,
                     'ztsecond2': a22,
                     'ldeltaall2': a23, 'mpolepath': a24, 'xtransformer': a25, 'ytransformer': a26, 'xcapacity': a27,
                     'ycapacity': a28, 'longitude': a29, 'latitude': a30, 'altitude': a31, 'brancher': a32,
                     'other1': a33,
                     'other2': a34})

        transz_parameter = transz_para
        trans_raw = transz_parameter
        print('transz_parameter=', transz_parameter)
    except:
        transz_parameter = []
        pass

    # ---------------无条件插入临时数据库
    conn = sqlite3.connect("rawdata.db3")
    cut = conn.cursor()
    try:
        cut.execute("""
                    create table transz (id INT not null, 
                                    tnode varchar(20),
                                    company      varchar(20),
                                    tname      varchar(200),
                                    tnumber      varchar(50),
                                    tcapacity      varchar(50),
                                    lname      varchar(50),
                                    tpath      varchar(200),
                                    lpath      varchar(200),
                                    polepath      varchar(200),
                                    rline      REAL,
                                    xline      REAL,
                                    zline      REAL,
                                    deltaline      REAL,
                                    allline_length      REAL,
                                    rt      REAL,
                                    xt      REAL,
                                    zt      REAL,
                                    deltazt      REAL,
                                    ztsecond      REAL,
                                    ldeltaall      REAL,
                                    ztsecond2      REAL,
                                    ldeltaall2      REAL,
                                    mpolepath      varchar(200),
                                    xtransformer      REAL,
                                    ytransformer       REAL,
                                    xcapacity      REAL,
                                    ycapacity      REAL,                                
                                    longitude      varchar(200),
                                    latitude      varchar(200),
                                    altitude     varchar(200),

                                    brancher     varchar(200),

                                    other1      varchar(200),
                                    other2      varchar(200)                                  
                                            ); """)

        for j in range(len(transz_parameter)):
            cut.execute(
                "INSERT INTO transz VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                [j, (transz_parameter[j]['tnode']), (transz_parameter[j]['company']),
                 (transz_parameter[j]['tname']),
                 (transz_parameter[j]['tnumber']), (transz_parameter[j]['tcapacity']),
                 (transz_parameter[j]['lname']),
                 (transz_parameter[j]['tpath']), (transz_parameter[j]['lpath']),
                 intput_polepath((str(transz_parameter[j]['polepath']))), (transz_parameter[j]['rline']),
                 (transz_parameter[j]['xline']), (transz_parameter[j]['zline']),
                 (transz_parameter[j]['deltaline']),
                 (transz_parameter[j]['allline_length']), (transz_parameter[j]['rt']),
                 (transz_parameter[j]['xt']), (transz_parameter[j]['zt']),
                 (transz_parameter[j]['deltazt']), (transz_parameter[j]['ztsecond']),
                 (transz_parameter[j]['ldeltaall']),
                 (transz_parameter[j]['ztsecond2']), (transz_parameter[j]['ldeltaall2']),

                 (transz_parameter[j]['mpolepath']),

                 (transz_parameter[j]['xtransformer']),
                 (transz_parameter[j]['ytransformer']),
                 (transz_parameter[j]['xcapacity']),
                 (transz_parameter[j]['ycapacity']),

                 (transz_parameter[j]['longitude']),
                 (transz_parameter[j]['latitude']), (transz_parameter[j]['altitude']),

                 intput_polepath((str(transz_parameter[j]['brancher']))),

                 (transz_parameter[j]['other1']), (transz_parameter[j]['other2'])])
            conn.commit()
        conn.commit()
    except:
        sql = "select * from transz"
        cut.execute(sql)
        total = len(cut.fetchall()) + 1
        fty = time.localtime()

        for j in range(len(transz_parameter)):
            transz_time = str(time.strftime("%Y%m%d%H%M%S", fty)) + str(transz_parameter[j]['id'])
            cut.execute(
                "INSERT INTO transz VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                [j, (transz_parameter[j]['tnode']), (transz_parameter[j]['company']),
                 (transz_parameter[j]['tname']),
                 (transz_parameter[j]['tnumber']), (transz_parameter[j]['tcapacity']),
                 (transz_parameter[j]['lname']),
                 (transz_parameter[j]['tpath']), (transz_parameter[j]['lpath']),
                 intput_polepath((str(transz_parameter[j]['polepath']))), (transz_parameter[j]['rline']),
                 (transz_parameter[j]['xline']), (transz_parameter[j]['zline']),
                 (transz_parameter[j]['deltaline']),
                 (transz_parameter[j]['allline_length']), (transz_parameter[j]['rt']),
                 (transz_parameter[j]['xt']), (transz_parameter[j]['zt']),
                 (transz_parameter[j]['deltazt']), (transz_parameter[j]['ztsecond']),
                 (transz_parameter[j]['ldeltaall']),
                 (transz_parameter[j]['ztsecond2']), (transz_parameter[j]['ldeltaall2']),
                 (transz_parameter[j]['mpolepath']),

                 (transz_parameter[j]['xtransformer']),
                 (transz_parameter[j]['ytransformer']),
                 (transz_parameter[j]['xcapacity']),
                 (transz_parameter[j]['ycapacity']),

                 (transz_parameter[j]['longitude']),
                 (transz_parameter[j]['latitude']), (transz_parameter[j]['altitude']),

                 intput_polepath((str(transz_parameter[j]['brancher']))),

                 (transz_parameter[j]['other1']), (transz_parameter[j]['other2'])])
            conn.commit()
    conn.close()

    # --------------------导入新数据End-------------------------------

    # ---------------------完好版本Start----------------------------
    def getlinepara(line_type):
        Line_para = []
        conn = sqlite3.connect("rawdata.db3")
        cur = conn.cursor()
        sql = "select  * from line_parameter where line_type='" + str(line_type) + "'"
        cur.execute(sql)
        total = cur.fetchall()
        conn.commit()
        conn.close()
        # print('len(total)=', len(total))
        Line_para.append([total[0][1], total[0][2], total[0][3], total[0][4], total[0][5], total[0][6]])
        '''
        if total:
            Line_para.append([total[0][1], total[0][2], total[0][3], total[0][4], total[0][5], total[0][6]])
        else:
            Line_para.append(['LGJ-16', 0.0, 0.0, 0.0, 0.0, 0.0])
        '''

        return Line_para[0]

    def gettrans_impedance(capa):
        # capacity = 'S11-200'
        # capacity = capa

        if capa:
            capacity = capa
        else:
            capacity = 'S0-0'

        # print('capacity=', capacity)
        capacity1 = capacity.split('-')
        # print('capacity1=', capacity1)
        capacity2 = capacity1[-1]
        capafact = int(capacity2)
        # print('capafact=', capafact)
        # print('type=', type(capafact))
        trans_para = []

        conn = sqlite3.connect("rawdata.db3")
        cur = conn.cursor()
        sql = "select  * from transformer where capacity='" + str(capacity2) + "'"
        cur.execute(sql)
        total = cur.fetchall()

        # print('len(total)=', len(total))
        conn.commit()
        conn.close()

        print('total=', total)

        if total:
            trans_para.append([total[0][4], total[0][5], total[0][6], total[0][7]])
        else:
            # return None
            trans_para.append([80, 0.01, 0.01, 0.01])
        trans_parameters = trans_para

        # print('trans_parameters=', trans_parameters)

        rt = float(trans_para[0][2])
        xt = float(trans_para[0][3])
        zt = np.sqrt(rt ** 2 + xt ** 2)
        deltazt = math.atan(xt / rt) * 180 / 3.14159  # print('transformer=', trans_parameters)

        return float(rt), float(xt), float(zt), float(deltazt)

    def transz(line_typepath, line_lengthpath):
        line_types = line_typepath.split('.')
        linez = []
        liner = []
        linex = []
        for iterm in range(len(line_types)):
            zline = getlinepara(line_types[iterm])[1]
            rline = getlinepara(line_types[iterm])[2]
            xline = getlinepara(line_types[iterm])[3]
            linez.append(zline)
            liner.append(rline)
            linex.append(xline)

        # print('liner=', liner)
        # print('linex=', linex)
        # print('linez=', linez)

        line_lengths = line_lengthpath.split('.')
        line_all_lengths = sum(list(map(float, line_lengths)))
        line_ldata = []
        liners = []
        linexs = []

        line_news_lengths = list(map(float, line_lengths))

        '''

        for k in range(len(line_lengths)):
            line_ldata.append(float(line_lengths[k]))
            liners.append(line_ldata[k] * liner[k])
            linexs.append(line_ldata[k] * linex[k])
        print('line_ldata=', line_ldata)
        '''

        # print('float(line_lengths=', line_news_lengths)
        # print('rall_new=', liner)

        # if len(line_news_lengths) == len(liner):
        if len(line_news_lengths) == len(liner):
            rall = np.dot(line_news_lengths, liner) / 1000
            xall = np.dot(line_news_lengths, linex) / 1000
            zall = np.sqrt(rall ** 2 + xall ** 2)

            if rall > 0.001:
                delta = math.atan(xall / rall) * 180 / 3.14159
            else:
                delta = 80.01
        else:
            rall = 0.01
            xall = 0.01
            zall = 0.01
            delta = 80.01

        # print('rall_new=', rall)
        # print('xall_new=', xall)

        # rall = sum(liners) / 1000
        # xall = sum(linexs) / 1000

        # print('rall=', rall)
        # print('xall=', xall)
        # print('zall=', zall)
        # print('delta=', delta)
        return rall, xall, zall, delta, line_all_lengths

    def uptransz(upstransz):
        transz_up = upstransz
        cond = sqlite3.connect("fault.db3")
        cud = cond.cursor()

        sqall = "select  * from transz"
        cud.execute(sqall)
        totalall = len(cud.fetchall())

        sqn = "select  * from transz where tname='" + str(transz_up[3]) + "' and lname='" + str(transz_up[6]) + "'"
        cud.execute(sqn)
        totaln = cud.fetchall()

        print('totaln=', totaln)

        if totaln:
            # pass
            '''
            cud.execute("delete from transz where id = (?)", [totaln[0][0]])
            cond.commit()
            '''
            ftx = time.localtime()
            transz_times = str(time.strftime("%Y%m%d%H%M%S", ftx)) + str(transz_up[0])
            cud.execute("replace  into  transz(id, tnode, company,tname, tnumber,tcapacity,lname, tpath, lpath, "
                        "polepath,rline, xline,zline, deltaline,allline_length, rt,xt, zt,deltazt,ztsecond,"
                        "ldeltaall,ztsecond2,ldeltaall2, mpolepath, xtransformer, ytransformer, xcapacity, ycapacity, "
                        "longitude,latitude,altitude, brancher, other1,other2 ) values  (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?, "
                        "?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        [totaln[0][0], (transz_up[1]), (transz_up[2]), (transz_up[3]), transz_times,
                         (transz_up[5]), (transz_up[6]), (transz_up[7]), (transz_up[8]), (transz_up[9]),
                         (transz_up[10]), (transz_up[11]), (transz_up[12]), (transz_up[13]), (transz_up[14]),
                         (transz_up[15]), (transz_up[16]), (transz_up[17]), (transz_up[18]), (transz_up[19]),
                         (transz_up[20]), (transz_up[21]), (transz_up[22]), (transz_up[23]), (transz_up[24]),
                         (transz_up[25]), (transz_up[26]), (transz_up[27]), (transz_up[28]), (transz_up[29]),
                         (transz_up[30]), (transz_up[31]), (transz_up[32]), (transz_up[33])])




        else:

            fty = time.localtime()

            transz_time = str(time.strftime("%Y%m%d%H%M%S", fty)) + str(transz_up[0])

            cud.execute("INSERT INTO transz VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,"
                        "?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        [totalall + 1, (transz_up[1]), (transz_up[2]), (transz_up[3]), transz_time,
                         (transz_up[5]), (transz_up[6]), (transz_up[7]), (transz_up[8]), (transz_up[9]),
                         (transz_up[10]), (transz_up[11]), (transz_up[12]), (transz_up[13]), (transz_up[14]),
                         (transz_up[15]), (transz_up[16]), (transz_up[17]), (transz_up[18]), (transz_up[19]),
                         (transz_up[20]), (transz_up[21]), (transz_up[22]), (transz_up[23]), (transz_up[24]),
                         (transz_up[25]), (transz_up[26]), (transz_up[27]), (transz_up[28]), (transz_up[29]),
                         (transz_up[30]), (transz_up[31]), (transz_up[32]), (transz_up[33])])

        cond.commit()
        cond.close()
        return True

    # ralls, xalls, zalls, deltas = transz(line_tpath, line_lpath)

    # print('ralls, xalls, zalls, deltas = ', ralls, xalls, zalls, deltas)

    cont = sqlite3.connect("rawdata.db3")
    cut = cont.cursor()
    # sqt = "select  * from transz where lname='" + str(Fault_line) + "'"
    sqt = "select  * from transz "
    cut.execute(sqt)
    transzalls = cut.fetchall()
    print('len(transalls)=', len(transzalls))
    # print('transzall=', transzalls)
    cont.commit()
    cont.close()

    # cont.commit()

    cont = sqlite3.connect("rawdata.db3")
    cut = cont.cursor()
    sqz = "select  * from transz "
    cut.execute(sqz)
    transzall = cut.fetchall()
    lentran = len(transzall)

    # transzall = cut.fetchone()
    # print('transzallooook=', transzall)
    # print('transzall[0][5]=', transzall[5])

    print('lentran =', lentran)

    # print(cut.fetchone())

    # print(cut.fetchone())
    cont.commit()
    cont.close()

    cont = sqlite3.connect("rawdata.db3")
    cut = cont.cursor()
    sqz = "select  * from transz "
    cut.execute(sqz)

    line_all = []

    for i in range(lentran):
        trans_single = transzall[i]
        # trans_single = cut.fetchone()
        line_all.append(trans_single[6])

        tid = trans_single[0]
        a6 = trans_single[5]
        a8 = trans_single[7]
        a9 = trans_single[8]

        print('tid=', tid)
        print('a6=', a6)
        print('a8=', a8)
        print('a9=', a9)

        rfct, xfact, zfact, dfact, l_allgth = transz(a8, a9)
        rtct, xtact, ztact, dtact = gettrans_impedance(a6)

        ztsecond = np.sqrt((rfct + rtct) ** 2 + (xfact + xtact) ** 2)
        delta_all = math.atan((xfact + xtact) / (rfct + rtct)) * 180 / 3.14159

        ztsecond2 = np.sqrt((rfct + 1.5 * rtct) ** 2 + (xfact + 1.5 * xtact) ** 2)
        delta_all2 = math.atan((xfact + 1.5 * xtact) / (rfct + 1.5 * rtct)) * 180 / 3.14159

        # 考虑低压侧单相对地短路数据
        trans_single = [trans_single[0], trans_single[1], trans_single[2], trans_single[3], trans_single[4],
                        trans_single[5], trans_single[6], trans_single[7], trans_single[8], trans_single[9],
                        rfct, xfact, zfact, dfact, l_allgth,
                        rtct, xtact, ztact, dtact, ztsecond,
                        delta_all, ztsecond2, delta_all2, trans_single[23], trans_single[24],
                        trans_single[25], trans_single[26], trans_single[27], trans_single[28], trans_single[29],
                        trans_single[30], trans_single[31], trans_single[32], trans_single[33]]

        print('trans_single=', trans_single)
        print('len(trans_single)=', len(trans_single))

        trans_correct = trans_single[9].replace('F', '').replace('H', '').replace('X', '').replace('Z', '').replace('Y', '').replace(
            'L', '').replace('&', '').replace('^', '')

        trans_correct1 = trans_correct.replace('.', '')


        if '-' in trans_single[5] and trans_correct1.isdecimal():

            uptransz(trans_single)
        else:
            win32api.MessageBox(0, "导入参数格式错误！请检查!", "提醒！", win32con.MB_OK)
        cont.commit()

    cont.close()

    # 通过excel表transz第二个表插入line数据表Start

    try:
        Line_ifs = []
        wb = load_workbook(r'inputdata\\transz.xlsx')
        ws = wb['line']

        for row_A in range(2, ws.max_row + 1, 1):
            a1 = ws.cell(row=row_A, column=1).value
            a2 = ws.cell(row=row_A, column=2).value
            a3 = ws.cell(row=row_A, column=3).value
            a4 = ws.cell(row=row_A, column=4).value
            a5 = ws.cell(row=row_A, column=5).value
            a6 = ws.cell(row=row_A, column=6).value
            a7 = ws.cell(row=row_A, column=7).value
            a8 = ws.cell(row=row_A, column=8).value
            a9 = ws.cell(row=row_A, column=9).value
            a10 = ws.cell(row=row_A, column=10).value
            a11 = ws.cell(row=row_A, column=11).value
            a12 = ws.cell(row=row_A, column=12).value
            a13 = ws.cell(row=row_A, column=13).value
            '''
            Line_ifs.append(
                {'县公司': a1, '厂站': a2, '线路名称': a3, '系统阻抗': a4, '系统阻抗电阻': a5, '系统阻抗电抗': a6, 'CT变比': a7,
                 '第一段线缆型号': a8, '第一段线缆长度': a9, '第二段线缆型号': a10, '第二段线缆长度': a11, '线路总长': a12,
                 '系统阻抗角': a13})
            '''
            Line_ifs.append(
                {'县公司': a1, '厂站': a2, '线路名称': a3, '系统阻抗': a4, '系统阻抗电阻': a5, '系统阻抗电抗': a6, 'CT变比': a7,
                 '第一段线缆型号': a8, '第一段线缆长度': a9, '第二段线缆型号': a10, '第二段线缆长度': a11, '线路总长': a12})
        line_raw = Line_ifs
        print('line_raw1=', line_raw)

        conn = sqlite3.connect("fault.db3")
        cur = conn.cursor()

        try:
            cur.execute("""
                                create table line (id INT not null, 
                                                    company varchar(20),
                                                    station varchar(20),
                                                    lname   varchar(20),
                                                    zs      REAL,
                                                    rs      REAL,
                                                    xs      REAL,
                                                    CT      INT,
                                                    line1_type   varchar(20),
                                                    line1_length      REAL,
                                                    line2_type   varchar(20),
                                                    line2_length      REAL,
                                                    line_all_length      REAL
                                                    ); """)
            for k in range(len(line_raw)):
                cur.execute("INSERT INTO line VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
                            [k, (line_raw[k]['县公司']), (line_raw[k]["厂站"]), (line_raw[k]["线路名称"]), (line_raw[k]['系统阻抗']),
                             (line_raw[k]['系统阻抗电阻']), (line_raw[k]['系统阻抗电抗']), (line_raw[k]['CT变比']),
                             (line_raw[k]['第一段线缆型号']),
                             (line_raw[k]['第一段线缆长度']), (line_raw[k]['第二段线缆型号']), (line_raw[k]['第二段线缆长度']),
                             (line_raw[k]['线路总长']), ])

        except:

            sql = "select * from line"
            cur.execute(sql)
            total = len(cur.fetchall()) + 1

            for k in range(len(line_raw)):

                sqline = "select  * from line where lname='" + str(line_raw[k]["线路名称"]) + "' and company='" + str(
                    line_raw[k]['县公司']) + "'"
                cur.execute(sqline)
                totaline = cur.fetchall()
                print('totaline=', totaline)
                print('\n')
                if totaline:

                    cur.execute("delete from line where lname='" + str(line_raw[k]["线路名称"]) + "' and company='" + str(
                        line_raw[k]['县公司']) + "'")
                    conn.commit()

                    cur.execute("replace  into  line(id, company,station, lname,zs,rs, xs, CT, "
                                "line1_type,line1_length, line2_type, line2_length,line_all_length) values  (?,?,?,?,?,?,"
                                "?,?,?,?,?,?,?)",
                                [totaline[0][0], (line_raw[k]['县公司']), (line_raw[k]["厂站"]), (line_raw[k]["线路名称"]),
                                 (line_raw[k]['系统阻抗']),
                                 (line_raw[k]['系统阻抗电阻']), (line_raw[k]['系统阻抗电抗']), (line_raw[k]['CT变比']),
                                 (line_raw[k]['第一段线缆型号']),
                                 (line_raw[k]['第一段线缆长度']), (line_raw[k]['第二段线缆型号']), (line_raw[k]['第二段线缆长度']),
                                 (line_raw[k]['线路总长']), ])

                else:
                    cur.execute("INSERT INTO line VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
                                [total + k - 1, (line_raw[k]['县公司']), (line_raw[k]["厂站"]), (line_raw[k]["线路名称"]),
                                 (line_raw[k]['系统阻抗']),
                                 (line_raw[k]['系统阻抗电阻']), (line_raw[k]['系统阻抗电抗']), (line_raw[k]['CT变比']),
                                 (line_raw[k]['第一段线缆型号']),
                                 (line_raw[k]['第一段线缆长度']), (line_raw[k]['第二段线缆型号']), (line_raw[k]['第二段线缆长度']),
                                 (line_raw[k]['线路总长']), ])
        print('line_raw2=', line_raw)
        conn.commit()
        conn.close()

    except:
        pass

    # -------------------------------------------

    # -----录入系统运行方式数据Start-----------------------------------

    # ----------------------buspath数据录入rawdata---------------------------------

    conn = sqlite3.connect("rawdata.db3")
    cut = conn.cursor()
    try:
        buspath_para = []
        wb = load_workbook(r'inputdata\\transz.xlsx')
        ws = wb['buspath']

        for row_A in range(2, ws.max_row + 1, 1):
            a1 = ws.cell(row=row_A, column=1).value
            a2 = ws.cell(row=row_A, column=2).value
            a3 = ws.cell(row=row_A, column=3).value
            a4 = ws.cell(row=row_A, column=4).value
            a5 = ws.cell(row=row_A, column=5).value
            a6 = ws.cell(row=row_A, column=6).value
            a7 = ws.cell(row=row_A, column=7).value
            a8 = ws.cell(row=row_A, column=8).value
            a9 = ws.cell(row=row_A, column=9).value
            a10 = ws.cell(row=row_A, column=10).value
            a11 = ws.cell(row=row_A, column=11).value
            a12 = ws.cell(row=row_A, column=12).value
            a13 = ws.cell(row=row_A, column=13).value
            a14 = ws.cell(row=row_A, column=14).value
            a15 = ws.cell(row=row_A, column=15).value
            a16 = ws.cell(row=row_A, column=16).value
            a17 = ws.cell(row=row_A, column=17).value
            a18 = ws.cell(row=row_A, column=18).value
            a19 = ws.cell(row=row_A, column=19).value
            a20 = ws.cell(row=row_A, column=20).value
            a21 = ws.cell(row=row_A, column=21).value
            a22 = ws.cell(row=row_A, column=22).value
            a23 = ws.cell(row=row_A, column=23).value
            a24 = ws.cell(row=row_A, column=24).value
            a25 = ws.cell(row=row_A, column=25).value
            a26 = ws.cell(row=row_A, column=26).value
            a27 = ws.cell(row=row_A, column=27).value

            buspath_para.append(
                {'id': a1, 'mode': a2, 'breakerpath': a3, 'busname': a4, 'busnumber': a5, 'bus_station': a6,
                 'tpath': a7,
                 'cpath': a8, 'dpath': a9, 'vpath': a10, 'lpath': a11, 'gpath': a12, 'rline': a13, 'xline': a14,
                 'zline': a15, 'deltaline': a16,
                 'allline_length': a17, 'rt': a18, 'xt': a19, 'zt': a20, 'deltazt': a21, 'rbus': a22, 'xbus': a23,
                 'zbus': a24,
                 'deltabus': a25, 'longitude': a26, 'latitude': a27})
        buspath_parameter = buspath_para
        buspath_raw = buspath_parameter
        print('buspath_parameter=', buspath_parameter)
        try:
            cut.execute("""
                create table buspath (id INT not null, 
                                mode varchar(50),
                                breakerpath      varchar(200),
                                busname      varchar(200),
                                busnumber      varchar(50),
                                bus_station      varchar(200),
                                tpath      varchar(200),
                                cpath      varchar(200),
                                dpath      varchar(200),
                                vpath      varchar(200),
                                lpath      varchar(200),
                                gpath      varchar(200),                            
                                rline      REAL,
                                xline      REAL,
                                zline      REAL,
                                deltaline      REAL,
                                allline_length      REAL,
                                rt      REAL,
                                xt      REAL,
                                zt      REAL,
                                deltazt      REAL,                            
                                rbus      REAL,
                                xbus      REAL,
                                zbus      REAL,
                                deltabus      REAL,
                                longitude      varchar(200),
                                latitude      varchar(200),
                                PRIMARY KEY(breakerpath, mode, busname, bus_station)    
                                ); """)
            for j in range(len(buspath_parameter)):
                cut.execute("INSERT INTO buspath VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                            [j, (buspath_parameter[j]['mode']), (buspath_parameter[j]['breakerpath']),
                             (buspath_parameter[j]['busname']), (buspath_parameter[j]['busnumber']),
                             (buspath_parameter[j]['bus_station']),
                             (buspath_parameter[j]['tpath']), (buspath_parameter[j]['cpath']),
                             (buspath_parameter[j]['dpath']),
                             (buspath_parameter[j]['vpath']), (buspath_parameter[j]['lpath']),
                             (buspath_parameter[j]['gpath']),
                             (buspath_parameter[j]['rline']), (buspath_parameter[j]['xline']),
                             (buspath_parameter[j]['zline']),
                             (buspath_parameter[j]['deltaline']), (buspath_parameter[j]['allline_length']),
                             (buspath_parameter[j]['rt']),
                             (buspath_parameter[j]['xt']), (buspath_parameter[j]['zt']),
                             (buspath_parameter[j]['deltazt']),
                             (buspath_parameter[j]['rbus']), (buspath_parameter[j]['xbus']),
                             (buspath_parameter[j]['zbus']),
                             (buspath_parameter[j]['deltabus']), (buspath_parameter[j]['longitude']),
                             (buspath_parameter[j]['latitude'])])
                # conn.commit()
        except:

            sql = "select * from buspath"
            cut.execute(sql)
            totalallbus = len(cut.fetchall()) + 1
            for j in range(len(buspath_parameter)):
                '''
                sqb = "select  * from buspath where busname='" + str(
                    buspath_parameter[j]['busname']) + "' and breakerpath='" + str(
                    buspath_parameter[j]['breakerpath']) + "' and bus_station='" + str(
                    buspath_parameter[j]['bus_station']) + "'"
                '''

                sqb = "select  * from buspath where busname='" + str(
                    buspath_parameter[j]['busname']) + "' and mode='" + str(
                    buspath_parameter[j]['mode']) + "' and bus_station='" + str(
                    buspath_parameter[j]['bus_station']) + "'"

                cut.execute(sqb)
                totalbus = cut.fetchall()
                print('totalbus=', totalbus)
                print('\n')
                if totalbus:
                    # 如有，则更新
                    # pass
                    ftbus = time.localtime()

                    buspath_time = str(time.strftime("%Y%m%d%H%M%S", ftbus)) + str(buspath_parameter[j]['id'])
                    cut.execute(
                        "replace  into  buspath(id, mode,breakerpath, busname,busnumber,bus_station, tpath, cpath, "
                        "dpath,vpath, lpath, gpath, rline, xline,zline, deltaline,allline_length, rt,xt, zt,"
                        "deltazt,rbus, xbus,zbus,deltabus, longitude,latitude ) values  (?,?,?,?,?,?,?,?,?,?,?,?,"
                        "?,?,?, ?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        [(buspath_parameter[j]['id']), (buspath_parameter[j]['mode']),
                         (buspath_parameter[j]['breakerpath']),
                         (buspath_parameter[j]['busname']), (buspath_parameter[j]['busnumber']),
                         (buspath_parameter[j]['bus_station']),
                         (buspath_parameter[j]['tpath']), (buspath_parameter[j]['cpath']),
                         (buspath_parameter[j]['dpath']),
                         (buspath_parameter[j]['vpath']), (buspath_parameter[j]['lpath']),
                         (buspath_parameter[j]['gpath']),
                         (buspath_parameter[j]['rline']), (buspath_parameter[j]['xline']),
                         (buspath_parameter[j]['zline']),
                         (buspath_parameter[j]['deltaline']), (buspath_parameter[j]['allline_length']),
                         (buspath_parameter[j]['rt']),
                         (buspath_parameter[j]['xt']), (buspath_parameter[j]['zt']),
                         (buspath_parameter[j]['deltazt']),
                         (buspath_parameter[j]['rbus']), (buspath_parameter[j]['xbus']),
                         (buspath_parameter[j]['zbus']),
                         (buspath_parameter[j]['deltabus']), (buspath_parameter[j]['longitude']),
                         (buspath_parameter[j]['latitude'])])
                    # conn.commit()



                else:
                    fty = time.localtime()

                    buspath_time = str(time.strftime("%Y%m%d%H%M%S", fty)) + str(buspath_parameter[j]['id'])

                    cut.execute("INSERT INTO buspath VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                                [totalallbus + j - 1, (buspath_parameter[j]['mode'].replace(',', '')),
                                 (buspath_parameter[j]['breakerpath']),
                                 (buspath_parameter[j]['busname']), buspath_time,
                                 (buspath_parameter[j]['bus_station']),
                                 (buspath_parameter[j]['tpath']), (buspath_parameter[j]['cpath']),
                                 (buspath_parameter[j]['dpath']),
                                 (buspath_parameter[j]['vpath']), (buspath_parameter[j]['lpath']),
                                 (buspath_parameter[j]['gpath']),
                                 (buspath_parameter[j]['rline']), (buspath_parameter[j]['xline']),
                                 (buspath_parameter[j]['zline']),
                                 (buspath_parameter[j]['deltaline']), (buspath_parameter[j]['allline_length']),
                                 (buspath_parameter[j]['rt']),
                                 (buspath_parameter[j]['xt']), (buspath_parameter[j]['zt']),
                                 (buspath_parameter[j]['deltazt']),
                                 (buspath_parameter[j]['rbus']), (buspath_parameter[j]['xbus']),
                                 (buspath_parameter[j]['zbus']),
                                 (buspath_parameter[j]['deltabus']), (buspath_parameter[j]['longitude']),
                                 (buspath_parameter[j]['latitude'])])
                    # conn.commit()
            # conn.close()

    except:
        pass

    conn.commit()
    conn.close()

    def getlinepara(line_type):
        # print('line_type=', line_type)
        Line_para = []
        wb1 = load_workbook(r'inputdata\\line_parameter.xlsx')
        ws1 = wb1['line_parameter']
        for row_A1 in range(2, ws1.max_row + 1, 1):
            a111 = ws1.cell(row=row_A1, column=1).value
            a21 = ws1.cell(row=row_A1, column=2).value
            a31 = ws1.cell(row=row_A1, column=3).value
            a41 = ws1.cell(row=row_A1, column=4).value
            a51 = ws1.cell(row=row_A1, column=5).value
            a61 = ws1.cell(row=row_A1, column=6).value
            # print('a111=', a111)
            if str(a111) == str(line_type):
                # print('a111=', a111)
                Line_para.append([a111, a21, a31, a41, a51, a61])
        line_parameters = Line_para
        # print('Line_para=', Line_para)
        wb1.close()
        return Line_para[0]

    def gettrans_impedance(capa, utt):
        # capacity = 'S11-200'
        capacity = capa
        uttrans = utt
        # print('capacity=', capacity)
        capacity1 = capacity.split('.')
        # print('capacity1=', capacity1)
        uttrans1 = uttrans.split('.')
        # print('uttrans1=', uttrans1)

        rtk = []
        xtk = []
        ktt = [230.0, 115.0, 37.0]
        for k in range(len(capacity1)):
            xtk.append((10.5 / 1.0) ** 2 * 0.1 * float(uttrans1[k]) / float(capacity1[k]))
            # rtk.append((10.5 / 1.0) ** 2 * 0.176327 * 0.1 * float(uttrans1[k]) / float(capacity1[k]))
            rtk.append((10.5 / 1.0) ** 2 * 0.03492 * 0.1 * float(uttrans1[k]) / float(capacity1[k]))
            # rtk.append((10.5 / 1.0) ** 2 * 0.085 * 0.1 * float(uttrans1[k]) / float(capacity1[k]))
        # print('xtk=', xtk)
        # print('rtk=', xtk)

        rtall = sum(rtk)
        xtall = sum(xtk)
        ztall = np.sqrt(rtall ** 2 + xtall ** 2)
        deltazt = math.atan(xtall / rtall) * 180 / 3.14159

        return float(rtall), float(xtall), float(ztall), float(deltazt)

    def transz_bus(vgpath, line_typepath, line_lengthpath):
        line_types = line_typepath.split('.')
        line_voltage = vgpath.split('.')
        # print('line_types=', line_types)
        # print('line_voltage=', line_voltage)
        linez = []
        liner = []
        linex = []
        for iterm in range(len(line_types)):
            zline = getlinepara(line_types[iterm])[1]
            rline = getlinepara(line_types[iterm])[2]
            xline = getlinepara(line_types[iterm])[3]
            linez.append(zline)
            liner.append(rline)
            linex.append(xline)
            # print('zline=', zline)
        # print('liner=', liner)
        # print('linex=', linex)
        # print('linez=', linez)

        line_lengths = line_lengthpath.split('.')
        line_all_lengths = sum(list(map(float, line_lengths)))
        line_ldata = []
        liners = []
        linexs = []
        kvv = []
        for g in range(len(line_voltage)):
            kvv.append(float(line_voltage[g]))

        for k in range(len(line_lengths)):
            line_ldata.append(float(line_lengths[k]))
            liners.append((10.5 / kvv[k]) ** 2 * line_ldata[k] * liner[k])
            linexs.append((10.5 / kvv[k]) ** 2 * line_ldata[k] * linex[k])
        # print('line_ldata=', line_ldata)

        rall = sum(liners) / 1000
        xall = sum(linexs) / 1000
        zall = np.sqrt(rall ** 2 + xall ** 2)

        if rall > 0.001:
            delta = math.atan(xall / rall) * 180 / 3.14159
        else:
            delta = 80.01

        # print('rall=', rall)
        # print('xall=', xall)
        # print('zall=', zall)
        # print('delta=', delta)
        return rall, xall, zall, delta, line_all_lengths

    # ralls, xalls, zalls, deltas = transz(line_tpath, line_lpath)

    # print('ralls, xalls, zalls, deltas = ', ralls, xalls, zalls, deltas)

    # ---------------------输出到excel数据中Start---------------

    wb = load_workbook('inputdata\\transz.xlsx')

    ws = wb['buspath']
    try:
        for row_A in range(2, ws.max_row + 1, 1):
            a1 = ws.cell(row=row_A, column=1).value
            a2 = ws.cell(row=row_A, column=2).value
            a3 = ws.cell(row=row_A, column=3).value
            a4 = ws.cell(row=row_A, column=4).value
            a5 = ws.cell(row=row_A, column=5).value
            a6 = ws.cell(row=row_A, column=6).value
            a7 = ws.cell(row=row_A, column=7).value
            a8 = ws.cell(row=row_A, column=8).value
            a9 = ws.cell(row=row_A, column=9).value
            a10 = ws.cell(row=row_A, column=10).value
            a11 = ws.cell(row=row_A, column=11).value
            a12 = ws.cell(row=row_A, column=12).value
            a13 = ws.cell(row=row_A, column=13).value
            a14 = ws.cell(row=row_A, column=14).value
            a15 = ws.cell(row=row_A, column=15).value
            a16 = ws.cell(row=row_A, column=16).value
            a17 = ws.cell(row=row_A, column=17).value
            a18 = ws.cell(row=row_A, column=18).value
            a19 = ws.cell(row=row_A, column=19).value
            a20 = ws.cell(row=row_A, column=20).value
            a21 = ws.cell(row=row_A, column=21).value
            a22 = ws.cell(row=row_A, column=22).value
            a23 = ws.cell(row=row_A, column=23).value
            a24 = ws.cell(row=row_A, column=24).value
            a25 = ws.cell(row=row_A, column=25).value
            a26 = ws.cell(row=row_A, column=26).value
            a27 = ws.cell(row=row_A, column=27).value

            print('a10=', a10)
            print('a11=', a11)
            print('a12=', a12)

            # a6 = 'LGJ-120.LGJ-95.LGJ-70'
            # a7 = '1250.750.100'
            rfct, xfact, zfact, dfact, l_allgth = transz_bus(str(a10), str(a11), str(a12))
            print('rfct, xfact, zfact, dfact, l_allgth=', rfct, xfact, zfact, dfact, l_allgth)
            try:

                rtct, xtact, ztact, dtact = gettrans_impedance(a8, a9)
                print('---------------------------------------------')
                ztsecond = np.sqrt((rfct + rtct) ** 2 + (xfact + xtact) ** 2)
                delta_all = math.atan((xfact + xtact) / (rfct + rtct)) * 180 / 3.14159

                print('rtct, xtact, ztact, dtact=', rtct, xtact, ztact, dtact)

                ws.cell(row=row_A, column=13).value = rfct
                ws.cell(row=row_A, column=14).value = xfact
                ws.cell(row=row_A, column=15).value = zfact
                ws.cell(row=row_A, column=16).value = dfact
                ws.cell(row=row_A, column=17).value = l_allgth

                ws.cell(row=row_A, column=18).value = rtct
                ws.cell(row=row_A, column=19).value = xtact
                ws.cell(row=row_A, column=20).value = ztact
                ws.cell(row=row_A, column=21).value = dtact

                ws.cell(row=row_A, column=22).value = rfct + rtct
                ws.cell(row=row_A, column=23).value = xfact + xtact

                ws.cell(row=row_A, column=24).value = ztsecond
                ws.cell(row=row_A, column=25).value = delta_all

                wb.save("inputdata\\transz.xlsx")


            except:

                rtct, xtact, ztact, dtact = 0.0, 0.0, 0.0, 0.0

                ztsecond = np.sqrt((rfct + rtct) ** 2 + (xfact + xtact) ** 2)
                delta_all = math.atan((xfact + xtact) / (rfct + rtct)) * 180 / 3.14159

                ws.cell(row=row_A, column=13).value = rfct
                ws.cell(row=row_A, column=14).value = xfact
                ws.cell(row=row_A, column=15).value = zfact
                ws.cell(row=row_A, column=16).value = dfact
                ws.cell(row=row_A, column=17).value = l_allgth

                ws.cell(row=row_A, column=18).value = rtct
                ws.cell(row=row_A, column=19).value = xtact
                ws.cell(row=row_A, column=20).value = ztact
                ws.cell(row=row_A, column=21).value = dtact

                ws.cell(row=row_A, column=22).value = rfct + rtct
                ws.cell(row=row_A, column=23).value = xfact + xtact

                ws.cell(row=row_A, column=24).value = ztsecond
                ws.cell(row=row_A, column=25).value = delta_all

                wb.save("inputdata\\transz.xlsx")
        wb.close()
        win32api.MessageBox(0, "导入成功!", "提醒！", win32con.MB_OK)
    except:
        wb.close()
        win32api.MessageBox(0, "导入参数格式错误！请检查!", "提醒！", win32con.MB_OK)

    # ---------------------输出到excel数据中End---------------

    # -------------结算结果插入数据库-----------------------

    def upbuspth(upstransz):
        transz_up = upstransz

        print('transz_up=', transz_up)

        cond = sqlite3.connect("fault.db3")
        cud = cond.cursor()

        sqall = "select  * from buspath"
        cud.execute(sqall)
        totalabus = len(cud.fetchall())

        # print('\n')
        print('totalabus=', totalabus)
        # print('\n')

        sqm = "select  * from buspath where busname='" + str(transz_up[3]) + "' and breakerpath='" + str(
            transz_up[2]) + "' and bus_station='" + str(transz_up[5]) + "'"
        cud.execute(sqm)
        total = cud.fetchall()
        print('\n')
        print('total=', total)
        print('\n')

        if total:
            # 如果有记录，则更新
            # pass

            ftbuspath = time.localtime()

            buspaths_time = str(time.strftime("%Y%m%d%H%M%S", ftbuspath)) + str(total[0][0])

            cud.execute("replace  into  buspath (id, mode,breakerpath, busname,busnumber,bus_station, tpath, cpath, "
                        "dpath,vpath, lpath, gpath, rline, xline,zline, deltaline,allline_length, rt,xt, zt,"
                        "deltazt,rbus, xbus,zbus,deltabus, longitude,latitude ) values  (?,?,?,?,?,?,?,?,?,?,?,?,"
                        "?,?,?, ?,?,?,?,?,?,?,?,?,?,?,?)",
                        [total[0][0], (transz_up[1]), (transz_up[2]), (transz_up[3]), buspaths_time, (transz_up[5]),
                         (transz_up[6]), (transz_up[7]), (transz_up[8]), (transz_up[9]), (transz_up[10]),
                         (transz_up[11]),
                         (transz_up[12]), (transz_up[13]), (transz_up[14]), (transz_up[15]), (transz_up[16]),
                         (transz_up[17]),
                         (transz_up[18]), (transz_up[19]), (transz_up[20]), (transz_up[21]), (transz_up[22]),
                         (transz_up[23]),
                         (transz_up[24]), (transz_up[25]), (transz_up[26])])

            cond.commit()
            cond.close()


        else:
            # 如果没有记录，则新增该记录

            print('--------------新增记录-------------------')

            fgbus = time.localtime()

            zbus_time = str(time.strftime("%Y%m%d%H%M%S", fgbus)) + str(transz_up[0])

            cud.execute("INSERT INTO buspath VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        [totalabus + 1, (transz_up[1]), (transz_up[2]), (transz_up[3]), zbus_time, (transz_up[5]),
                         (transz_up[6]), (transz_up[7]), (transz_up[8]), (transz_up[9]), (transz_up[10]),
                         (transz_up[11]),
                         (transz_up[12]), (transz_up[13]), (transz_up[14]), (transz_up[15]), (transz_up[16]),
                         (transz_up[17]),
                         (transz_up[18]), (transz_up[19]), (transz_up[20]), (transz_up[21]), (transz_up[22]),
                         (transz_up[23]),
                         (transz_up[24]), (transz_up[25]), (transz_up[26])])

            cond.commit()
            cond.close()
        return True

    cont = sqlite3.connect("rawdata.db3")
    cut = cont.cursor()

    sqt = "select  * from buspath "
    cut.execute(sqt)
    transzalls = cut.fetchall()
    print('len(transalls)=', len(transzalls))

    cont.commit()
    cont.close()

    cont = sqlite3.connect("rawdata.db3")
    cut = cont.cursor()
    sqz = "select  * from buspath "
    cut.execute(sqz)
    transzall = cut.fetchall()
    lentran = len(transzall)

    print('lentran =', lentran)

    cont.commit()
    cont.close()

    cont = sqlite3.connect("rawdata.db3")
    cut = cont.cursor()
    sqz = "select  * from  buspath "
    cut.execute(sqz)

    try:
        for i in range(lentran):
            buspath_single = cut.fetchone()

            bid = buspath_single[0]
            a7 = buspath_single[7]
            a8 = buspath_single[8]
            a9 = buspath_single[9]
            a10 = buspath_single[10]
            a11 = buspath_single[11]

            # print('tid=', bid)
            # print('a6=', a9)
            # print('a8=', a10)
            # print('a9=', a11)

            # transz(vgpath, line_typepath, line_lengthpath):
            rfct, xfact, zfact, dfact, l_allgth = transz_bus(str(a9), str(a10), str(a11))
            # print('rfct, xfact, zfact, dfact, l_allgth=', rfct, xfact, zfact, dfact, l_allgth)
            try:

                rtct, xtact, ztact, dtact = gettrans_impedance(a7, a8)

                ztsecond = np.sqrt((rfct + rtct) ** 2 + (xfact + xtact) ** 2)
                delta_all = math.atan((xfact + xtact) / (rfct + rtct)) * 180 / 3.14159
                print('---------------------------------------------')

                # gettrans_impedance(capa, utt):
                # print('rtct, xtact, ztact, dtact=', rtct, xtact, ztact, dtact)

                buspath_single = [buspath_single[0], buspath_single[1], buspath_single[2], buspath_single[3],
                                  buspath_single[4], buspath_single[5], buspath_single[6], buspath_single[7],
                                  buspath_single[8], buspath_single[9], buspath_single[10], buspath_single[11],
                                  rfct, xfact, zfact, dfact, l_allgth, rtct, xtact, ztact, dtact,
                                  rfct + rtct, xfact + xtact, ztsecond, delta_all, buspath_single[25],
                                  buspath_single[26]]

                print('trans_single=', buspath_single)
                upbuspth(buspath_single)


            except:

                rtct, xtact, ztact, dtact = 0.0, 0.0, 0.0, 0.0
                ztsecond = np.sqrt((rfct + rtct) ** 2 + (xfact + xtact) ** 2)
                delta_all = math.atan((xfact + xtact) / (rfct + rtct)) * 180 / 3.14159

                buspath_single = [buspath_single[0], buspath_single[1], buspath_single[2], buspath_single[3],
                                  buspath_single[4], buspath_single[5], buspath_single[6], buspath_single[7],
                                  buspath_single[8], buspath_single[9], buspath_single[10], buspath_single[11],
                                  rfct, xfact, zfact, dfact, l_allgth, rtct, xtact, ztact, dtact,
                                  rfct + rtct, xfact + xtact, ztsecond, delta_all, buspath_single[25],
                                  buspath_single[26]]

                print('trans_single=', buspath_single)
                upbuspth(buspath_single)
        win32api.MessageBox(0, "导入成功!", "提醒！", win32con.MB_OK)




    except:

        win32api.MessageBox(0, "导入参数格式错误！请检查!", "提醒！", win32con.MB_OK)

    cont.commit()

    cont.close()

    conw = sqlite3.connect("rawdata.db3")
    cuw = conw.cursor()
    cuw.execute("delete from buspath")
    conw.commit()
    conw.close()

    # 录入系统运行方式数据End------------------------------------------

    print('line_all =', line_all)

    line_list = sorted(list(set(line_all)))
    print('line_list =', line_list)

    for k in range(len(line_list)):
        try:
            polemodify(line_list[k])
        except:
            pass


    '''
    cont = sqlite3.connect("rawdata.db3")
    cut = cont.cursor()
    sqv = "delete   from transz "
    cut.execute(sqv)
    cont.commit()
    cont.close()
    '''

    '''
    conu = sqlite3.connect("rawdata.db3")
    cuu = cont.cursor()
    squ = "selete   from transz "
    cuu.execute(squ)
    transzact = cuu.fetchall()
    conu.commit()
    conu.close()
    
    if transzact:
        win32api.MessageBox(0, "导入失败!", "提醒！", win32con.MB_OK)
    else:
        win32api.MessageBox(0, "导入成功!", "提醒！", win32con.MB_OK)
    
    '''
    # ----------------out_transz-----------------



    # -------------------------------------------

    return True
